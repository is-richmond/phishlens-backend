"""
Bulk Generation Router

Endpoints for bulk generation from Excel files.
"""

from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, status, File, UploadFile, Form, Query
from sqlalchemy.orm import Session

from app.core.deps import get_db, get_current_user
from app.models.user import User
from app.models.bulk_generation import BulkGeneration, BulkGenerationResult
from app.models.campaign import Campaign
from app.models.distribution import Distribution
from app.schemas.bulk_generation import (
    BulkGenerationUploadResponse,
    FieldMappingRequest,
    BulkGenerationStartRequest,
    BulkGenerationProgressResponse,
    BulkGenerationResultsResponse,
    BulkGenerationResultItem,
    BulkGenerationDetailResponse,
)
from app.schemas.distribution import (
    BulkDistributeRequest,
    BulkDistributeResponse,
)
from app.services.bulk_generation_service import bulk_generation_service
from app.services.excel_service import excel_service
from app.services.distribution_service import distribution_service
from app.core.logging import get_logger
import asyncio

logger = get_logger("bulk_generations_router")

router = APIRouter(prefix="/bulk-generations", tags=["Bulk Generations"])


@router.post("/upload", response_model=BulkGenerationUploadResponse)
async def upload_bulk_file(
    file: UploadFile = File(...),
    title: str = Form(...),
    scenario_id: UUID = Form(...),
    template_id: Optional[UUID] = Form(None),
    description: Optional[str] = Form(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Upload Excel file for bulk generation.
    
    Returns file preview with auto-detected field mapping.
    """
    try:
        # Read file
        content = await file.read()
        if not content:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File is empty"
            )

        logger.info(f"Uploading bulk file: {file.filename}")

        # Create bulk generation campaign
        bulk_gen = bulk_generation_service.create_bulk_generation(
            db=db,
            user_id=current_user.id,
            title=title,
            original_filename=file.filename or "upload.xlsx",
            file_data=content,
            scenario_id=scenario_id,
            template_id=template_id,
            description=description,
        )

        # Get preview data
        headers, rows, _ = excel_service.parse_excel_file(content, max_rows=3)
        preview_rows = rows[:3]

        return BulkGenerationUploadResponse(
            bulk_generation_id=bulk_gen.id,
            total_rows=bulk_gen.total_rows,
            column_headers=headers,
            auto_mapping=bulk_gen.field_mapping,
            preview_rows=preview_rows,
        )

    except Exception as e:
        logger.error(f"Error uploading bulk file: {e}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error processing file: {str(e)}"
        )


@router.get("/{bulk_generation_id}", response_model=BulkGenerationDetailResponse)
async def get_bulk_generation(
    bulk_generation_id: UUID,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get details of a bulk generation campaign."""
    bulk_gen = db.query(BulkGeneration).filter(
        BulkGeneration.id == bulk_generation_id,
        BulkGeneration.user_id == current_user.id,
    ).first()

    if not bulk_gen:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Bulk generation not found"
        )

    # Parse Excel file to get column headers and preview rows
    try:
        headers, rows, _ = excel_service.parse_excel_file(bulk_gen.file_data, max_rows=3)
        preview_rows = rows[:3]
    except Exception as e:
        logger.warning(f"Could not parse Excel file for preview: {e}")
        headers = []
        preview_rows = []

    # Create response with additional fields
    return BulkGenerationDetailResponse(
        id=bulk_gen.id,
        title=bulk_gen.title,
        original_filename=bulk_gen.original_filename,
        description=bulk_gen.description,
        status=bulk_gen.status,
        scenario_id=bulk_gen.scenario_id,
        template_id=bulk_gen.template_id,
        field_mapping=bulk_gen.field_mapping,
        column_headers=headers,
        preview_rows=preview_rows,
        total_rows=bulk_gen.total_rows,
        generated_count=bulk_gen.generated_count,
        failed_count=bulk_gen.failed_count,
        temperature=str(bulk_gen.temperature),
        max_tokens=bulk_gen.max_tokens,
        model_variant=bulk_gen.model_variant or "gemini-2.5-flash-lite",
        created_at=bulk_gen.created_at,
        updated_at=bulk_gen.updated_at,
    )


@router.patch("/{bulk_generation_id}/field-mapping")
async def update_field_mapping(
    bulk_generation_id: UUID,
    request: FieldMappingRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Update field mapping for bulk generation."""
    bulk_gen = db.query(BulkGeneration).filter(
        BulkGeneration.id == bulk_generation_id,
        BulkGeneration.user_id == current_user.id,
    ).first()

    if not bulk_gen:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Bulk generation not found"
        )

    try:
        bulk_gen = bulk_generation_service.update_field_mapping(
            db=db,
            bulk_gen_id=bulk_generation_id,
            field_mapping=request.field_mapping,
        )
        return bulk_gen
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )


@router.post("/{bulk_generation_id}/generate")
async def start_bulk_generation(
    bulk_generation_id: UUID,
    request: BulkGenerationStartRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Start bulk generation process.
    
    Starts async background task for generation.
    """
    bulk_gen = db.query(BulkGeneration).filter(
        BulkGeneration.id == bulk_generation_id,
        BulkGeneration.user_id == current_user.id,
    ).first()

    if not bulk_gen:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Bulk generation not found"
        )

    if bulk_gen.status not in ["uploaded", "mapped"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot start generation: status is {bulk_gen.status}"
        )

    logger.info(f"Starting bulk generation {bulk_generation_id}")

    # Start async task (fire-and-forget)
    # In production, use Celery or similar task queue
    asyncio.create_task(
        bulk_generation_service.process_bulk_generation_async(
            db=db,
            bulk_gen_id=bulk_generation_id,
            temperature=request.temperature or 0.7,
            max_tokens=request.max_tokens or 1024,
            model_variant=request.model_variant or "gemini-2.5-flash-lite",
        )
    )

    return {
        "status": "queued",
        "message": "Bulk generation started in background",
        "bulk_generation_id": bulk_generation_id,
    }


@router.get("/{bulk_generation_id}/progress", response_model=BulkGenerationProgressResponse)
async def get_progress(
    bulk_generation_id: UUID,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get current progress of bulk generation."""
    bulk_gen = db.query(BulkGeneration).filter(
        BulkGeneration.id == bulk_generation_id,
        BulkGeneration.user_id == current_user.id,
    ).first()

    if not bulk_gen:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Bulk generation not found"
        )

    progress = bulk_generation_service.get_bulk_generation_progress(db, bulk_generation_id)
    
    # Log progress updates (helps debugging)
    if bulk_gen.status == "processing":
        logger.info(
            f"Progress check: {bulk_generation_id} | status={progress['status']} | "
            f"{progress['generated_count']}/{progress['total_rows']} (~{progress['progress_percent']}%)"
        )
    
    return BulkGenerationProgressResponse(
        bulk_generation_id=bulk_generation_id,
        **progress
    )


@router.get("/{bulk_generation_id}/results", response_model=BulkGenerationResultsResponse)
async def get_results(
    bulk_generation_id: UUID,
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get paginated results for bulk generation."""
    bulk_gen = db.query(BulkGeneration).filter(
        BulkGeneration.id == bulk_generation_id,
        BulkGeneration.user_id == current_user.id,
    ).first()

    if not bulk_gen:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Bulk generation not found"
        )

    data = bulk_generation_service.get_bulk_generation_results(
        db=db,
        bulk_gen_id=bulk_generation_id,
        page=page,
        per_page=per_page,
    )

    # Convert to response model
    results = [
        BulkGenerationResultItem(
            id=r.id,
            row_index=r.row_index,
            input_data=r.input_data,
            generated_subject=r.generated_subject,
            generated_message=r.generated_message,
            status=r.status,
            error_message=r.error_message,
            created_at=r.created_at,
        )
        for r in data["results"]
    ]

    return BulkGenerationResultsResponse(
        results=results,
        pagination=data["pagination"],
    )


@router.get("/{bulk_generation_id}/results/export")
async def export_results(
    bulk_generation_id: UUID,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Export bulk generation results as Excel file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    bulk_gen = db.query(BulkGeneration).filter(
        BulkGeneration.id == bulk_generation_id,
        BulkGeneration.user_id == current_user.id,
    ).first()

    if not bulk_gen:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Bulk generation not found"
        )

    # Get all results
    all_results = db.query(BulkGenerationResult).filter(
        BulkGenerationResult.bulk_generation_id == bulk_generation_id
    ).order_by(BulkGenerationResult.row_index).all()

    # Create workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Results"

    # Add headers
    headers = ["Row", "Status", "Input Data", "Generated Subject", "Generated Message", "Error"]
    for col, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data rows
    for row_idx, result in enumerate(all_results, start=2):
        worksheet.cell(row=row_idx, column=1).value = result.row_index
        worksheet.cell(row=row_idx, column=2).value = result.status
        worksheet.cell(row=row_idx, column=3).value = str(result.input_data)
        worksheet.cell(row=row_idx, column=4).value = result.generated_subject
        worksheet.cell(row=row_idx, column=5).value = result.generated_message
        worksheet.cell(row=row_idx, column=6).value = result.error_message

    # Adjust column widths
    worksheet.column_dimensions['A'].width = 8
    worksheet.column_dimensions['B'].width = 12
    worksheet.column_dimensions['C'].width = 25
    worksheet.column_dimensions['D'].width = 30
    worksheet.column_dimensions['E'].width = 50
    worksheet.column_dimensions['F'].width = 30

    # Save to bytes
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=bulk_results_{bulk_generation_id}.xlsx"}
    )


@router.delete("/{bulk_generation_id}")
async def delete_bulk_generation(
    bulk_generation_id: UUID,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Delete a bulk generation campaign and all its results."""
    bulk_gen = db.query(BulkGeneration).filter(
        BulkGeneration.id == bulk_generation_id,
        BulkGeneration.user_id == current_user.id,
    ).first()

    if not bulk_gen:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Bulk generation not found"
        )

    db.delete(bulk_gen)
    db.commit()
    logger.info(f"Deleted bulk generation {bulk_generation_id}")

    return {"message": "Bulk generation deleted"}


@router.post("/{bulk_generation_id}/distribute", response_model=BulkDistributeResponse)
async def distribute_bulk_messages(
    bulk_generation_id: UUID,
    request: BulkDistributeRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Create distributions (email records) from bulk generation results.
    
    This endpoint:
    1. Creates a Campaign from the bulk generation
    2. Creates a Distribution record for each generated message
    3. Optionally sends immediately or queues for background sending
    
    Args:
        bulk_generation_id: ID of the bulk generation
        request: Distribution request with optional campaign name
        current_user: Current authenticated user
        db: Database session
        
    Returns:
        Summary of created distributions and campaign
    """
    try:
        # Verify bulk generation exists and belongs to user
        bulk_gen = db.query(BulkGeneration).filter(
            BulkGeneration.id == bulk_generation_id,
            BulkGeneration.user_id == current_user.id,
        ).first()
        
        if not bulk_gen:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Bulk generation not found"
            )
        
        if bulk_gen.status != "completed":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Bulk generation must be completed before distributing. Current status: {bulk_gen.status}"
            )
        
        logger.info(f"Creating distributions for bulk generation {bulk_generation_id}")
        
        # Create campaign
        campaign_name = request.campaign_name or f"Distribution: {bulk_gen.title}"
        campaign = Campaign(
            name=campaign_name,
            description=f"Bulk distribution from: {bulk_gen.title}",
            user_id=current_user.id,
        )
        db.add(campaign)
        db.flush()  # Get campaign.id without committing yet
        
        # Create distributions from bulk generation results
        distributions_created, error = distribution_service.create_distributions_from_bulk_generation(
            db=db,
            bulk_generation_id=bulk_generation_id,
            campaign_id=campaign.id,
        )
        
        if error:
            db.rollback()
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=error
            )
        
        if distributions_created == 0:
            db.rollback()
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No distributions could be created from the bulk generation results"
            )
        
        # Commit campaign creation
        db.commit()
        
        # Optionally send immediately
        if request.send_immediately:
            logger.info(f"Starting immediate distribution for {distributions_created} recipients")
            asyncio.create_task(
                distribution_service.send_pending_distributions(
                    db=db,
                    batch_size=int(distributions_created),
                )
            )
            status_msg = "sending_in_background"
        else:
            status_msg = "queued_for_sending"
        
        # Get distribution statistics
        stats = distribution_service.get_distribution_stats(db, campaign.id)
        
        logger.info(
            f"Distribution created: {distributions_created} recipients, "
            f"campaign_id={campaign.id}, send_immediately={request.send_immediately}"
        )
        
        return BulkDistributeResponse(
            campaign_id=campaign.id,
            distributions_created=distributions_created,
            distributions_pending=stats["pending"],
            status=status_msg,
            message=f"Created {distributions_created} distributions. "
                    f"Campaign '{campaign_name}' is ready for sending.",
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error distributing bulk messages: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error creating distributions: {str(e)}"
        )
