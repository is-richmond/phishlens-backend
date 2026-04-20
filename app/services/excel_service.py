"""
Excel Service — Parsing and field mapping for bulk generation

Handles uploading, parsing, and field mapping for Excel files.
"""

import re
from typing import Dict, List, Tuple, Any
from io import BytesIO

try:
    import openpyxl
except ImportError:
    openpyxl = None

from app.core.logging import get_logger

logger = get_logger("excel_service")

# Common placeholder patterns to auto-detect
PLACEHOLDER_PATTERNS = {
    r"(?i)(ФИО|full.?name|name|имя|фамилия)": "[TARGET_NAME]",
    r"(?i)(email|почта|e-mail|электронная)": "[TARGET_EMAIL]",
    r"(?i)(должность|position|title|роль)": "[TARGET_DEPARTMENT]",
    r"(?i)(компания|company|организация|org|организацию)": "[COMPANY_NAME]",
    r"(?i)(отдел|department|dept)": "[TARGET_DEPARTMENT]",
    r"(?i)(телефон|phone|tel)": "[PHONE_NUMBER]",
    r"(?i)(служба доставки|delivery.?service|сервис)": "[DELIVERY_SERVICE]",
    r"(?i)(номер отслеживания|tracking.?number|трек)": "[TRACKING_NUMBER]",
}


class ExcelService:
    """Service for parsing and managing Excel files."""

    @staticmethod
    def parse_excel_file(
        file_data: bytes, max_rows: int = 100
    ) -> Tuple[List[str], List[Dict[str, Any]], int]:
        """Parse Excel file and extract headers and rows.

        Args:
            file_data: Raw file bytes
            max_rows: Maximum rows to process

        Returns:
            Tuple of (column_headers, rows_data, total_row_count)
        """
        if not openpyxl:
            logger.error("openpyxl not installed")
            raise RuntimeError("Excel support not installed. Install openpyxl package.")

        try:
            workbook = openpyxl.load_workbook(BytesIO(file_data))
            worksheet = workbook.active

            logger.info(f"Parsing Excel file: {worksheet.title}")

            # Extract headers from first row
            headers: List[str] = []
            for cell in worksheet[1]:
                if cell.value is not None:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append("")

            logger.info(f"Found {len(headers)} columns: {headers}")

            # Extract data rows
            rows: List[Dict[str, Any]] = []
            row_count = 0

            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), start=2):
                if row_idx - 1 > max_rows:
                    logger.warning(f"Stopping at {max_rows} rows (file has more)")
                    break

                row_data = {}
                for col_idx, (header, cell) in enumerate(zip(headers, row)):
                    if header:  # Skip empty header columns
                        row_data[header] = cell.value

                # Only include rows that have at least one non-empty value
                if any(v is not None for v in row_data.values()):
                    rows.append(row_data)
                    row_count += 1

            total_rows = worksheet.max_row - 1  # Exclude header
            logger.info(f"Parsed {len(rows)} rows (total in file: {total_rows})")

            return headers, rows, total_rows

        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}")
            raise

    @staticmethod
    def auto_detect_field_mapping(headers: List[str]) -> Dict[int, str]:
        """Auto-detect field mapping from column headers.

        Tries to match headers to common placeholder patterns.

        Args:
            headers: Column headers from Excel

        Returns:
            Mapping of column_index -> placeholder_name
        """
        mapping: Dict[int, str] = {}

        for col_idx, header in enumerate(headers):
            if not header:
                continue

            matched = False
            for pattern, placeholder in PLACEHOLDER_PATTERNS.items():
                if re.search(pattern, header):
                    mapping[str(col_idx)] = placeholder
                    logger.debug(f"Auto-mapped '{header}' -> {placeholder}")
                    matched = True
                    break

            if not matched:
                # If no pattern matches, suggest as potential custom field
                logger.debug(f"No auto-match for column '{header}'")

        logger.info(f"Auto-detected mapping: {mapping}")
        return mapping

    @staticmethod
    def apply_field_replacements(
        row_data: Dict[str, Any],
        headers: List[str],
        field_mapping: Dict[str, str],
    ) -> Dict[str, str]:
        """Apply field replacements to a row.

        Takes row data and applies field mapping to create replacement dict.

        Args:
            row_data: Single row from Excel ({"ФИО": "Иван", "Email": "ivan@company.com", ...})
            headers: Column headers
            field_mapping: Mapping of column_index -> placeholder

        Returns:
            Replacements dict: {placeholder -> value}
        """
        replacements: Dict[str, str] = {}

        for col_idx_str, placeholder in field_mapping.items():
            try:
                col_idx = int(col_idx_str)
                if col_idx < len(headers):
                    header = headers[col_idx]
                    value = row_data.get(header, "")

                    # Convert to string and sanitize
                    if value is not None:
                        value_str = str(value).strip()
                        replacements[placeholder] = value_str
                        logger.debug(f"{placeholder} = {value_str}")

            except (ValueError, IndexError) as e:
                logger.warning(f"Error applying mapping for column {col_idx_str}: {e}")

        return replacements

    @staticmethod
    def validate_mapping(
        field_mapping: Dict[str, str], headers: List[str]
    ) -> Tuple[bool, List[str]]:
        """Validate field mapping against headers.

        Args:
            field_mapping: Proposed mapping
            headers: Column headers

        Returns:
            Tuple of (is_valid, error_messages)
        """
        errors: List[str] = []

        for col_idx_str, placeholder in field_mapping.items():
            try:
                col_idx = int(col_idx_str)
                if col_idx >= len(headers):
                    errors.append(f"Column index {col_idx} out of range (max: {len(headers) - 1})")

                if not placeholder.startswith("[") or not placeholder.endswith("]"):
                    errors.append(f"Invalid placeholder format: {placeholder} (must be [PLACEHOLDER])")

            except ValueError:
                errors.append(f"Invalid column index: {col_idx_str}")

        return len(errors) == 0, errors


excel_service = ExcelService()
